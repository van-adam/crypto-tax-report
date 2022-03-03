import datetime as d

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


# tuple: (date: datetime.date, quantity: float, price: float); sorted by date ascending
sells = []
buys = []

# sells = [
#     (d.date(2021, 4, 16), 0.13541820, 35),
#     (d.date(2021, 4, 21), 0.11454640, 25),
#     (d.date(2021, 4, 22), 0.13420283, 30),
#     (d.date(2021, 5, 3), 0.36125077, 88.09),
# ]
#
# buys = [
#     (d.date(2020, 5, 27), 0.18586880, 7.44),
#     (d.date(2020, 5, 27), 0.08721041, 3.49),
#     (d.date(2020, 6, 4), 0.27289865, 11.71),
#     (d.date(2020, 6, 10), 0.27456618, 11.22),
# ]


# sells = [
#     (d.date(2021, 5, 1), 3.0, 100),
#     (d.date(2021, 6, 1), 1.0, 50),
#     (d.date(2021, 7, 1), 2.0, 120),
#     (d.date(2021, 8, 1), 2.0, 150),
# ]
#
# buys = [
#     (d.date(2020, 5, 1), 1.0, 10),
#     (d.date(2020, 6, 1), 2.0, 25),
#     (d.date(2020, 7, 1), 2.0, 45),
#     (d.date(2020, 8, 1), 3.0, 60),
# ]


def import_transactions_from_file(file_path: str, sheet_name: str, destination_list: list) -> None:
    """
    Imports a list of transactions from the passed XLSX file to the passed destination list.

    :param file_path: the file path to the XLSX file containing the transactions
    :param sheet_name: the name of the sheet containing the transactions
    :param destination_list: the list to write the transactions to
    """
    workbook = load_workbook(file_path)
    sheet: Worksheet = workbook[sheet_name]

    if len(destination_list) != 0:
        destination_list.clear()

    for row in range(2, sheet.max_row + 1):
        date = sheet.cell(row, 1).value
        date = date.date()
        quantity = sheet.cell(row, 2).value
        price = sheet.cell(row, 3).value

        destination_list.append((date, quantity, price))

    print("Imported transactions from file '{}' from sheet '{}'".format(file_path, sheet_name))


def print_buys() -> None:
    """
    Convenience method to print transactions.buys.
    """
    print_transactions(buys, "BUYS")


def print_sells():
    """
    Convenience method to print transactions.buys.
    """
    print_transactions(sells, "SELLS")


def to_string(t: tuple) -> str:
    """
    Formats a transaction tuple as a String.

    :param t: the transaction tuple
    :return: a formatted String based on the tuple
    """
    date: d.date = t[0]
    quantity = t[1]
    price = t[2]

    return "{:<10}: {:<10} :: {:<7.2f}".format(date.strftime('%d-%m-%Y'), quantity, price)


def print_transactions(transactions: list, label: str) -> None:
    """
    Prints the passed list of transactions to the console.

    :param transactions: the transactions list
    :param label: the label above the output
    :return:
    """
    print(label, "*" * (50 - len(label)))

    for t in transactions:
        print(to_string(t))

    print("=" * 50)
