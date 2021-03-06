import datetime
import logging
import datetime as d

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

import van.adam.config as c

log = logging.getLogger()

# tuple: (date: datetime.date, quantity: float, price: float); sorted by date ascending
sells = []
buys = []

# sells = [
#     (d.date(2021, 5, 1), 3.0, 100),
#     (d.date(2021, 6, 1), 1.0, 50),
#     (d.date(2021, 7, 1), 2.0, 120),
#     (d.date(2021, 8, 1), 2.0, 150),
# ]
#
# buys = [
#     (d.date(2020, 5, 1), 1.0, 10),
#     (d.date(2020, 6, 1), 2.0, 25),
#     (d.date(2020, 7, 1), 2.0, 45),
#     (d.date(2020, 8, 1), 3.0, 60),
# ]


def import_transactions_from_file(file_path: str, sheet_name: str, destination_list: list) -> None:
    """
    Imports a list of transactions from the passed XLSX file to the passed destination list.

    :param file_path: the file path to the XLSX file containing the transactions
    :param sheet_name: the name of the sheet containing the transactions
    :param destination_list: the list to write the transactions to
    """
    workbook = load_workbook(file_path)
    sheet: Worksheet = workbook[sheet_name]

    if len(destination_list) != 0:
        destination_list.clear()

    for row in range(2, sheet.max_row + 1):
        date = sheet.cell(row, 1).value
        date = date.date()
        quantity = sheet.cell(row, 2).value
        price = sheet.cell(row, 3).value

        destination_list.append((date, quantity, price))

    log.info("Imported transactions from file '{}' from sheet '{}'".format(file_path, sheet_name))


def print_buys() -> None:
    """
    Convenience method to print transactions.buys.
    """
    print_transactions(buys, "BUYS")


def print_sells() -> None:
    """
    Convenience method to print transactions.buys.
    """
    print_transactions(sells, "SELLS")


def to_string(t: tuple) -> str:
    """
    Formats a transaction tuple as a String.

    :param t: the transaction tuple
    :return: a formatted String based on the tuple
    """
    date: d.date = t[0]
    quantity = t[1]
    price = t[2]

    return "{:<10}: {:<10} :: {:<7.2f}".format(date.strftime('%d-%m-%Y'), quantity, price)


def print_transactions(transactions: list, label: str) -> None:
    """
    Prints the passed list of transactions to the console.

    :param transactions: the transactions list
    :param label: the label above the output
    :return:
    """
    line_length = 95
    log.info("{}{}".format(label, "*" * (line_length - len(label))))

    for t in transactions:
        log.info(to_string(t))

    log.info("=" * line_length)


def is_taxable(buy_date: datetime.date, sell_date: datetime.date) -> bool:
    """
    Determines whether a sell transaction is taxable based on the buy date and sell date. A sale is taxable whenever
    the difference between buy date and sell date is smaller than 360 days.

    :param buy_date: the buy date
    :param sell_date: the sell date
    :return: true if the sale is taxable
    """
    diff: datetime.timedelta = sell_date - buy_date
    if diff.days < c.TAXFREE_TIMEDELTA:
        taxable = True
    else:
        taxable = False

    return taxable
